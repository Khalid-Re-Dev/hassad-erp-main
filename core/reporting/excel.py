"""
Excel report generation with RTL support for Arabic.
Uses openpyxl for XLSX generation.
"""

from datetime import datetime
from decimal import Decimal
from typing import Dict, Any, List
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .schemas import ReportType


class ExcelReportGenerator:
    """Generate Excel reports with proper formatting and RTL support."""

    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self, report_type: ReportType, data: Dict[str, Any], language: str) -> tuple[bytes, str]:
        """
        Generate Excel file for report.
        
        Args:
            report_type: Type of report
            data: Report data dictionary
            language: 'ar' or 'en'
            
        Returns:
            Tuple of (file_bytes, filename)
        """
        wb = Workbook()
        ws = wb.active
        
        # Set RTL for Arabic
        if language == 'ar':
            ws.sheet_view.rightToLeft = True
        
        # Route to specific report generator
        if report_type == ReportType.TRIAL_BALANCE:
            self._generate_trial_balance(ws, data, language)
        elif report_type == ReportType.BALANCE_SHEET:
            self._generate_balance_sheet(ws, data, language)
        elif report_type == ReportType.INCOME_STATEMENT:
            self._generate_income_statement(ws, data, language)
        elif report_type == ReportType.SALES_BY_DAY:
            self._generate_sales_by_day(ws, data, language)
        elif report_type == ReportType.SALES_BY_PRODUCT:
            self._generate_sales_by_product(ws, data, language)
        elif report_type == ReportType.PURCHASES_SUMMARY:
            self._generate_purchases_summary(ws, data, language)
        elif report_type == ReportType.STOCK_MOVEMENT:
            self._generate_stock_movement(ws, data, language)
        elif report_type == ReportType.INVENTORY_VALUATION:
            self._generate_inventory_valuation(ws, data, language)
        elif report_type == ReportType.SUPPLIER_LEDGER:
            self._generate_supplier_ledger(ws, data, language)
        else:
            self._generate_generic(ws, data, language)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        
        filename = f"{report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return content, filename

    def _generate_trial_balance(self, ws, data: Dict[str, Any], language: str):
        """Generate trial balance Excel sheet."""
        ws.title = "Trial Balance" if language == 'en' else "ميزان المراجعة"
        
        # Title
        title = data['title_ar'] if language == 'ar' else data['title_en']
        ws['A1'] = title
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # Date
        ws['A2'] = f"As of: {data['as_of_date']}" if language == 'en' else f"كما في: {data['as_of_date']}"
        ws.merge_cells('A2:E2')
        
        # Headers
        row = 4
        if language == 'ar':
            headers = ['رمز الحساب', 'اسم الحساب', 'النوع', 'مدين', 'دائن']
        else:
            headers = ['Account Code', 'Account Name', 'Type', 'Debit', 'Credit']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data rows
        row += 1
        for item in data['rows']:
            ws.cell(row=row, column=1, value=item['account_code'])
            ws.cell(row=row, column=2, value=item['account_name_ar'] if language == 'ar' else item['account_name_en'])
            ws.cell(row=row, column=3, value=item['account_type'])
            ws.cell(row=row, column=4, value=float(item['debit']))
            ws.cell(row=row, column=5, value=float(item['credit']))
            
            # Format numbers
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            
            # Borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Totals
        ws.cell(row=row, column=3, value='Total' if language == 'en' else 'المجموع').font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(data['total_debits'])).font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(data['total_credits'])).font = Font(bold=True)
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        
        # Balance check
        row += 1
        balance_text = "Balanced ✓" if data['balanced'] else "NOT BALANCED ✗"
        balance_text_ar = "متوازن ✓" if data['balanced'] else "غير متوازن ✗"
        ws.cell(row=row, column=3, value=balance_text_ar if language == 'ar' else balance_text)
        ws.cell(row=row, column=3).font = Font(bold=True, color="00FF00" if data['balanced'] else "FF0000")
        
        # Auto-size columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _generate_balance_sheet(self, ws, data: Dict[str, Any], language: str):
        """Generate balance sheet Excel sheet."""
        ws.title = "Balance Sheet" if language == 'en' else "الميزانية العمومية"
        
        row = 1
        # Title
        title = data['title_ar'] if language == 'ar' else data['title_en']
        ws.cell(row=row, column=1, value=title).font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1, value=f"As of: {data['as_of_date']}" if language == 'en' else f"كما في: {data['as_of_date']}")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        
        # Assets section
        ws.cell(row=row, column=1, value='ASSETS' if language == 'en' else 'الأصول').font = Font(bold=True, size=12)
        row += 1
        
        for asset in data['assets']:
            ws.cell(row=row, column=1, value=asset['code'])
            ws.cell(row=row, column=2, value=asset['name_ar'] if language == 'ar' else asset['name_en'])
            ws.cell(row=row, column=3, value=float(asset['balance']))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        ws.cell(row=row, column=2, value='Total Assets' if language == 'en' else 'إجمالي الأصول').font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(data['total_assets'])).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        row += 2
        
        # Liabilities section
        ws.cell(row=row, column=1, value='LIABILITIES' if language == 'en' else 'الخصوم').font = Font(bold=True, size=12)
        row += 1
        
        for liability in data['liabilities']:
            ws.cell(row=row, column=1, value=liability['code'])
            ws.cell(row=row, column=2, value=liability['name_ar'] if language == 'ar' else liability['name_en'])
            ws.cell(row=row, column=3, value=float(liability['balance']))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        ws.cell(row=row, column=2, value='Total Liabilities' if language == 'en' else 'إجمالي الخصوم').font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(data['total_liabilities'])).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        row += 2
        
        # Equity section
        ws.cell(row=row, column=1, value='EQUITY' if language == 'en' else 'حقوق الملكية').font = Font(bold=True, size=12)
        row += 1
        
        for equity in data['equity']:
            ws.cell(row=row, column=1, value=equity['code'])
            ws.cell(row=row, column=2, value=equity['name_ar'] if language == 'ar' else equity['name_en'])
            ws.cell(row=row, column=3, value=float(equity['balance']))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        ws.cell(row=row, column=2, value='Total Equity' if language == 'en' else 'إجمالي حقوق الملكية').font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(data['total_equity'])).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        # Auto-size columns
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _generate_income_statement(self, ws, data: Dict[str, Any], language: str):
        """Generate income statement Excel sheet."""
        ws.title = "Income Statement" if language == 'en' else "قائمة الدخل"
        
        row = 1
        title = data['title_ar'] if language == 'ar' else data['title_en']
        ws.cell(row=row, column=1, value=title).font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        period = f"From {data['date_from']} to {data['date_to']}" if language == 'en' else f"من {data['date_from']} إلى {data['date_to']}"
        ws.cell(row=row, column=1, value=period)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        
        # Revenue section
        ws.cell(row=row, column=1, value='REVENUE' if language == 'en' else 'الإيرادات').font = Font(bold=True, size=12)
        row += 1
        
        for rev in data['revenue']:
            ws.cell(row=row, column=1, value=rev['code'])
            ws.cell(row=row, column=2, value=rev['name_ar'] if language == 'ar' else rev['name_en'])
            ws.cell(row=row, column=3, value=float(rev['balance']))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        ws.cell(row=row, column=2, value='Total Revenue' if language == 'en' else 'إجمالي الإيرادات').font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(data['total_revenue'])).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        row += 2
        
        # Expenses section
        ws.cell(row=row, column=1, value='EXPENSES' if language == 'en' else 'المصروفات').font = Font(bold=True, size=12)
        row += 1
        
        for exp in data['expenses']:
            ws.cell(row=row, column=1, value=exp['code'])
            ws.cell(row=row, column=2, value=exp['name_ar'] if language == 'ar' else exp['name_en'])
            ws.cell(row=row, column=3, value=float(exp['balance']))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        ws.cell(row=row, column=2, value='Total Expenses' if language == 'en' else 'إجمالي المصروفات').font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(data['total_expenses'])).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        row += 2
        
        # Net income
        ws.cell(row=row, column=2, value='NET INCOME' if language == 'en' else 'صافي الدخل').font = Font(bold=True, size=12)
        ws.cell(row=row, column=3, value=float(data['net_income'])).font = Font(bold=True, size=12)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        # Color code net income
        if data['net_income'] >= 0:
            ws.cell(row=row, column=3).font = Font(bold=True, size=12, color="00FF00")
        else:
            ws.cell(row=row, column=3).font = Font(bold=True, size=12, color="FF0000")
        
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _generate_sales_by_day(self, ws, data: Dict[str, Any], language: str):
        """Generate sales by day report."""
        ws.title = "Sales by Day" if language == 'en' else "المبيعات حسب اليوم"
        self._generate_tabular_report(ws, data, language, [
            ('date', 'Date' if language == 'en' else 'التاريخ'),
            ('transaction_count', 'Transactions' if language == 'en' else 'العمليات'),
            ('total_sales', 'Total Sales' if language == 'en' else 'إجمالي المبيعات'),
            ('total_tax', 'Tax' if language == 'en' else 'الضريبة')
        ])

    def _generate_sales_by_product(self, ws, data: Dict[str, Any], language: str):
        """Generate sales by product report."""
        ws.title = "Sales by Product" if language == 'en' else "المبيعات حسب المنتج"
        self._generate_tabular_report(ws, data, language, [
            ('sku', 'SKU' if language == 'en' else 'رمز المنتج'),
            ('product_name_ar' if language == 'ar' else 'product_name_en', 'Product' if language == 'en' else 'المنتج'),
            ('quantity', 'Quantity' if language == 'en' else 'الكمية'),
            ('total_sales', 'Total Sales' if language == 'en' else 'إجمالي المبيعات')
        ])

    def _generate_purchases_summary(self, ws, data: Dict[str, Any], language: str):
        """Generate purchases summary report."""
        ws.title = "Purchases" if language == 'en' else "المشتريات"
        self._generate_tabular_report(ws, data, language, [
            ('supplier_name', 'Supplier' if language == 'en' else 'المورد'),
            ('invoice_count', 'Invoices' if language == 'en' else 'الفواتير'),
            ('total_amount', 'Total Amount' if language == 'en' else 'المبلغ الإجمالي'),
            ('total_tax', 'Tax' if language == 'en' else 'الضريبة')
        ])

    def _generate_stock_movement(self, ws, data: Dict[str, Any], language: str):
        """Generate stock movement report."""
        ws.title = "Stock Movement" if language == 'en' else "حركة المخزون"
        self._generate_tabular_report(ws, data, language, [
            ('date', 'Date' if language == 'en' else 'التاريخ'),
            ('product_sku', 'SKU' if language == 'en' else 'رمز المنتج'),
            ('product_name_ar' if language == 'ar' else 'product_name_en', 'Product' if language == 'en' else 'المنتج'),
            ('movement_type', 'Type' if language == 'en' else 'النوع'),
            ('quantity', 'Quantity' if language == 'en' else 'الكمية'),
            ('unit_cost', 'Unit Cost' if language == 'en' else 'تكلفة الوحدة'),
            ('total_cost', 'Total Cost' if language == 'en' else 'التكلفة الإجمالية')
        ])

    def _generate_inventory_valuation(self, ws, data: Dict[str, Any], language: str):
        """Generate inventory valuation report."""
        ws.title = "Inventory Valuation" if language == 'en' else "تقييم المخزون"
        self._generate_tabular_report(ws, data, language, [
            ('sku', 'SKU' if language == 'en' else 'رمز المنتج'),
            ('product_name_ar' if language == 'ar' else 'product_name_en', 'Product' if language == 'en' else 'المنتج'),
            ('quantity', 'Quantity' if language == 'en' else 'الكمية'),
            ('avg_cost', 'Avg Cost' if language == 'en' else 'متوسط التكلفة'),
            ('total_value', 'Total Value' if language == 'en' else 'القيمة الإجمالية')
        ])

    def _generate_supplier_ledger(self, ws, data: Dict[str, Any], language: str):
        """Generate supplier ledger report."""
        ws.title = "Supplier Ledger" if language == 'en' else "دفتر الموردين"
        self._generate_tabular_report(ws, data, language, [
            ('supplier_name', 'Supplier' if language == 'en' else 'المورد'),
            ('tax_id', 'Tax ID' if language == 'en' else 'الرقم الضريبي'),
            ('total_invoiced', 'Invoiced' if language == 'en' else 'المفوتر'),
            ('total_paid', 'Paid' if language == 'en' else 'المدفوع'),
            ('outstanding', 'Outstanding' if language == 'en' else 'المستحق')
        ])

    def _generate_tabular_report(self, ws, data: Dict[str, Any], language: str, columns: List[tuple]):
        """Generic tabular report generator."""
        row = 1
        
        # Title
        title = data.get('title_ar' if language == 'ar' else 'title_en', 'Report')
        ws.cell(row=row, column=1, value=title).font = self.title_font
        ws.merge_cells(f'A{row}:{get_column_letter(len(columns))}{row}')
        
        row += 1
        
        # Date range if available
        if 'date_from' in data and 'date_to' in data:
            period = f"From {data['date_from']} to {data['date_to']}" if language == 'en' else f"من {data['date_from']} إلى {data['date_to']}"
            ws.cell(row=row, column=1, value=period)
            ws.merge_cells(f'A{row}:{get_column_letter(len(columns))}{row}')
            row += 1
        elif 'as_of_date' in data:
            ws.cell(row=row, column=1, value=f"As of: {data['as_of_date']}" if language == 'en' else f"كما في: {data['as_of_date']}")
            ws.merge_cells(f'A{row}:{get_column_letter(len(columns))}{row}')
            row += 1
        
        row += 1
        
        # Headers
        for col_idx, (field, header) in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        row += 1
        
        # Data rows
        for item in data.get('rows', []):
            for col_idx, (field, _) in enumerate(columns, start=1):
                value = item.get(field, '')
                if isinstance(value, Decimal):
                    value = float(value)
                ws.cell(row=row, column=col_idx, value=value)
                
                # Format numbers
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
                
                ws.cell(row=row, column=col_idx).border = self.border
            
            row += 1
        
        # Grand total if available
        if 'grand_total' in data:
            ws.cell(row=row, column=len(columns)-1, value='Total' if language == 'en' else 'المجموع').font = Font(bold=True)
            ws.cell(row=row, column=len(columns), value=float(data['grand_total'])).font = Font(bold=True)
            ws.cell(row=row, column=len(columns)).number_format = '#,##0.00'
        elif 'total_value' in data:
            ws.cell(row=row, column=len(columns)-1, value='Total' if language == 'en' else 'المجموع').font = Font(bold=True)
            ws.cell(row=row, column=len(columns), value=float(data['total_value'])).font = Font(bold=True)
            ws.cell(row=row, column=len(columns)).number_format = '#,##0.00'
        elif 'total_outstanding' in data:
            ws.cell(row=row, column=len(columns)-1, value='Total' if language == 'en' else 'المجموع').font = Font(bold=True)
            ws.cell(row=row, column=len(columns), value=float(data['total_outstanding'])).font = Font(bold=True)
            ws.cell(row=row, column=len(columns)).number_format = '#,##0.00'
        
        # Auto-size columns
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _generate_generic(self, ws, data: Dict[str, Any], language: str):
        """Generic report generator for unknown report types."""
        ws.title = "Report"
        
        row = 1
        ws.cell(row=row, column=1, value="Generic Report").font = self.title_font
        row += 2
        
        # Just dump the data as key-value pairs
        for key, value in data.items():
            if isinstance(value, (list, dict)):
                continue
            ws.cell(row=row, column=1, value=str(key))
            ws.cell(row=row, column=2, value=str(value))
            row += 1
